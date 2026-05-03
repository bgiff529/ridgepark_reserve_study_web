from __future__ import annotations

from io import BytesIO
import os
from pathlib import Path
import shutil
import subprocess
import sys
import tempfile

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from shiny import App, Inputs, Outputs, Session, reactive, render, ui

APP_DIR = Path(__file__).resolve().parent
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from reserve_plots import build_all_plots
from reserve_study_web_adapter import (
    ASSESSMENT_INPUT_COLUMNS,
    COMPONENT_INPUT_COLUMNS,
    DEFAULT_PROJECTION_YEARS,
    DEFAULT_UNITS,
    APP_ROOT,
    PROJECT_ROOT,
    coerce_assumptions_frame,
    prepare_assessment_input,
    prepare_components_input,
    run_reserve_study,
)


DEFAULT_VARIANT = os.environ.get("DEFAULT_VARIANT", "2026_joint_buget_maint")
SOURCE_ROOT = PROJECT_ROOT / DEFAULT_VARIANT / "source_data"
COMPONENT_SOURCE = SOURCE_ROOT / "component_list_v2.csv"
ASSUMPTIONS_SOURCE = SOURCE_ROOT / "assumptions.csv"
ASSESSMENT_SOURCE = SOURCE_ROOT / "assessment_contributions.csv"

APP_PASSWORD = os.environ.get("APP_PASSWORD", "")
ALLOW_NO_PASSWORD = os.environ.get("ALLOW_NO_PASSWORD", "").lower() in {"1", "true", "yes"}

COMPONENT_LABELS = {
    "category": "Category",
    "subcategory": "Subcategory",
    "component": "Component",
    "method": "Method",
    "cost": "Cost",
    "cost_units": "Cost Units",
    "quantity": "Quantity",
    "quantity_units": "Quantity Units",
    "useful_life": "Useful Life",
    "remaining_useful_life": "Remaining Useful Life",
    "notes": "Notes",
}
COMPONENT_REVERSE_LABELS = {label: column for column, label in COMPONENT_LABELS.items()}
NUMERIC_COMPONENT_COLUMNS = {"cost", "quantity", "useful_life"}


def load_default_inputs() -> dict[str, pd.DataFrame]:
    return {
        "components": prepare_components_input(pd.read_csv(COMPONENT_SOURCE)),
        "assumptions": coerce_assumptions_frame(pd.read_csv(ASSUMPTIONS_SOURCE)),
        "assessments": prepare_assessment_input(pd.read_csv(ASSESSMENT_SOURCE)),
    }


def display_components(df: pd.DataFrame) -> pd.DataFrame:
    out = df[COMPONENT_INPUT_COLUMNS].copy()
    return out.rename(columns=COMPONENT_LABELS)


def raw_components(df: pd.DataFrame) -> pd.DataFrame:
    out = df.rename(columns=COMPONENT_REVERSE_LABELS).copy()
    return prepare_components_input(out)


def patch_frame(
    state: reactive.Value[pd.DataFrame],
    patch: render.CellPatch,
    *,
    display_labels: dict[str, str] | None = None,
    numeric_columns: set[str] | None = None,
) -> object:
    df = state.get().copy()
    column_index = int(patch["column_index"])
    column = df.columns[column_index]
    raw_column = display_labels.get(column, column) if display_labels else column
    value = patch["value"]

    if numeric_columns and raw_column in numeric_columns:
        value = pd.to_numeric(value, errors="coerce")
        value = 0.0 if pd.isna(value) else float(value)
    else:
        value = "" if value is None else str(value)

    df.iat[int(patch["row_index"]), column_index] = value
    state.set(df)
    return value


def recalculate_workbook_if_possible(path: Path) -> Path:
    executable = shutil.which("soffice") or shutil.which("libreoffice")
    if executable is None:
        return path

    temp_dir = Path(tempfile.mkdtemp(prefix="reserve-study-upload-"))
    try:
        subprocess.run(
            [
                executable,
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(temp_dir),
                str(path),
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            check=False,
            timeout=45,
        )
        recalculated = temp_dir / f"{path.stem}.xlsx"
        return recalculated if recalculated.exists() else path
    except Exception:
        return path


def read_component_upload(path: str, filename: str) -> pd.DataFrame:
    upload_path = Path(path)
    suffix = Path(filename).suffix.lower()

    if suffix == ".csv":
        return prepare_components_input(pd.read_csv(upload_path))

    if suffix in {".xlsx", ".xlsm", ".xls", ".xlscsv"}:
        spreadsheet_path = recalculate_workbook_if_possible(upload_path)
        if suffix == ".xls":
            frame = pd.read_excel(spreadsheet_path, sheet_name="Component List")
        else:
            workbook = load_workbook(spreadsheet_path, data_only=True, read_only=True)
            if "Component List" not in workbook.sheetnames:
                raise ValueError('Spreadsheet imports must include a sheet named "Component List".')
            rows = list(workbook["Component List"].iter_rows(values_only=True))
            header_index = next(
                (
                    index
                    for index, row in enumerate(rows)
                    if row and str(row[0]).strip().lower() in {"category", "component #", "#"}
                ),
                None,
            )
            if header_index is None:
                raise ValueError('Could not find the header row on the "Component List" sheet.')
            header = [str(value).strip() if value is not None else "" for value in rows[header_index]]
            values = rows[header_index + 1 :]
            frame = pd.DataFrame(values, columns=header)
        return prepare_components_input(frame.dropna(how="all"))

    raise ValueError("Upload a CSV, XLSX, XLSM, XLS, or XLSCSV file.")


def component_template_bytes() -> bytes:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    component_list = workbook.create_sheet("Component List")

    title_fill = PatternFill("solid", fgColor="FFF347")
    input_fill = PatternFill("solid", fgColor="FFF9B1")
    header_fill = PatternFill("solid", fgColor="B7D7F0")
    bold = Font(bold=True)

    instructions["B2"] = "Ridge Park Reserve Study"
    instructions["B2"].font = Font(bold=True, size=18)
    instructions["B4"] = "Master Template for Component Schedule Import"
    instructions["B6"] = "Enter components on the Component List sheet. You may add extra tabs."
    instructions["B7"] = 'The importer reads only the sheet named "Component List".'
    instructions["B8"] = "If Component List contains formulas, save the workbook after recalculation before uploading."
    instructions["B10"] = "Required columns"
    instructions["B10"].fill = title_fill
    instructions["B10"].font = bold
    for row_num, column in enumerate(COMPONENT_INPUT_COLUMNS, start=11):
        instructions.cell(row=row_num, column=2, value=COMPONENT_LABELS[column])

    headers = [COMPONENT_LABELS[column] for column in COMPONENT_INPUT_COLUMNS]
    for column_num, header in enumerate(headers, start=1):
        cell = component_list.cell(row=1, column=column_num, value=header)
        cell.fill = header_fill
        cell.font = bold
        component_list.column_dimensions[cell.column_letter].width = max(14, len(header) + 3)

    sample = {
        "category": "Asphalt",
        "subcategory": "Top Coat",
        "component": "Asphalt Seal Coat",
        "method": "Repeating",
        "cost": 0.40,
        "cost_units": "$/SF",
        "quantity": 100000,
        "quantity_units": "SF",
        "useful_life": 5,
        "remaining_useful_life": "14:06",
        "notes": "Source page: 62",
    }
    for column_num, column in enumerate(COMPONENT_INPUT_COLUMNS, start=1):
        component_list.cell(row=2, column=column_num, value=sample[column]).fill = input_fill
    for row in component_list.iter_rows(min_row=3, max_row=250, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.fill = input_fill

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def money(value: object) -> str:
    try:
        return f"${float(value):,.0f}"
    except Exception:
        return str(value)


def app_shell() -> ui.Tag:
    return ui.div(
        ui.div(
            ui.div("Ridge Park", class_="rp-brand-mark"),
            ui.div(
                ui.div("Ridge Park Reserve Study", class_="rp-study-title"),
                ui.div("Reserve planning workspace", class_="rp-study-subtitle"),
            ),
            class_="rp-hero",
        ),
        ui.navset_tab(
            ui.nav_panel(
                "Components",
                ui.div(
                    ui.div(
                        ui.input_action_button("open_upload", "Upload"),
                        ui.download_button("download_components", "Download CSV"),
                        ui.input_action_button("reset_components", "Reset"),
                        ui.input_action_button("run_study", "Run Study", class_="rp-primary"),
                        class_="rp-toolbar",
                    ),
                    ui.output_data_frame("components_grid"),
                    class_="rp-panel",
                ),
            ),
            ui.nav_panel(
                "Assumptions",
                ui.div(
                    ui.div(
                        ui.input_action_button("run_study_assumptions", "Run Study", class_="rp-primary"),
                        class_="rp-toolbar",
                    ),
                    ui.output_data_frame("assumptions_grid"),
                    class_="rp-panel",
                ),
            ),
            ui.nav_panel(
                "Recommendations",
                ui.div(ui.output_table("statement_table"), class_="rp-panel"),
            ),
            ui.nav_panel(
                "Funding Plan Override",
                ui.div(
                    ui.p("Funding optimization controls will come later; this Shiny pass keeps the component workflow focused."),
                    class_="rp-panel rp-empty",
                ),
            ),
            ui.nav_panel(
                "Tables and Charts",
                ui.div(
                    ui.output_ui("results_summary"),
                    ui.output_plot("plot_0"),
                    ui.output_plot("plot_1"),
                    ui.output_plot("plot_2"),
                    ui.output_plot("plot_3"),
                    ui.output_plot("plot_4"),
                    ui.output_plot("plot_5"),
                    class_="rp-panel rp-plots",
                ),
            ),
            id="workspace_tab",
        ),
        ui.div(ui.download_button("download_template", "THIS TEMPLATE"), style="display:none;"),
        class_="rp-app",
    )


app_ui = ui.page_fillable(
    ui.head_content(
        ui.tags.meta(name="viewport", content="width=device-width, initial-scale=1"),
        ui.tags.style(
            """
            :root {
                --rp-navy: #1d2f3f;
                --rp-blue: #b8dfec;
                --rp-action: #6fd34f;
                --rp-border: #dfe7ec;
                --rp-bg: #f4f6f8;
                --rp-text: #2f3b44;
            }
            body {
                background: var(--rp-bg);
                color: var(--rp-text);
                font-family: Arial, Helvetica, sans-serif;
            }
            .container-fluid {
                padding: 0;
            }
            .rp-app {
                max-width: 1180px;
                margin: 28px auto 64px;
            }
            .rp-hero {
                height: 150px;
                background: var(--rp-navy);
                border: 1px solid #101b25;
                display: flex;
                align-items: end;
                padding: 0 0 18px 140px;
                position: relative;
            }
            .rp-brand-mark {
                position: absolute;
                left: 16px;
                bottom: -34px;
                width: 104px;
                height: 104px;
                border: 4px solid #fff;
                background: #f1f3f5;
                color: #789;
                display: flex;
                align-items: center;
                justify-content: center;
                text-align: center;
                font-weight: 700;
                box-shadow: 0 1px 4px rgb(0 0 0 / 18%);
            }
            .rp-study-title {
                color: #fff;
                font-size: 21px;
                font-weight: 700;
                line-height: 1.1;
            }
            .rp-study-subtitle {
                color: #fff;
                font-size: 12px;
                font-weight: 700;
            }
            .nav-tabs {
                margin-top: 28px;
                border: 0;
                display: flex;
                justify-content: space-between;
            }
            .nav-tabs .nav-link {
                border: 0;
                border-bottom: 2px solid transparent;
                border-radius: 0;
                color: #6d7983;
                font-size: 12px;
                font-weight: 700;
                letter-spacing: .02em;
                text-transform: uppercase;
                padding: 12px 28px;
            }
            .nav-tabs .nav-link.active {
                background: transparent;
                color: #357a9a;
                border-bottom-color: #8cc8dc;
            }
            .tab-content {
                margin-top: 16px;
            }
            .rp-panel {
                background: #fff;
                border: 1px solid var(--rp-border);
                min-height: 560px;
                padding: 18px;
            }
            .rp-toolbar {
                display: flex;
                justify-content: flex-end;
                gap: 8px;
                margin-bottom: 10px;
            }
            .rp-toolbar .btn {
                background: var(--rp-action);
                border-color: #5cc345;
                color: #24482a;
                font-size: 12px;
                font-weight: 700;
                padding: 4px 11px;
            }
            .rp-toolbar .btn.rp-primary {
                color: #17351b;
            }
            .rp-empty {
                color: #6d7983;
                font-size: 13px;
            }
            .rp-plots img, .rp-plots svg {
                max-width: 100%;
            }
            .modal-content {
                background: #0f141b;
                color: #fff;
                border-radius: 16px;
            }
            .modal-header, .modal-footer {
                border-color: #252b34;
            }
            .modal-title {
                font-size: 28px;
                font-weight: 800;
            }
            .rp-upload-copy {
                font-size: 18px;
                line-height: 1.55;
                margin-bottom: 18px;
            }
            .rp-upload-copy a {
                color: #fff;
                font-weight: 800;
                text-decoration: underline;
            }
            @media (max-width: 760px) {
                .rp-app {
                    margin: 12px;
                }
                .rp-hero {
                    padding-left: 18px;
                    align-items: center;
                }
                .rp-brand-mark {
                    display: none;
                }
                .nav-tabs {
                    overflow-x: auto;
                    justify-content: flex-start;
                }
                .nav-tabs .nav-link {
                    padding: 10px 14px;
                    white-space: nowrap;
                }
                .rp-toolbar {
                    justify-content: flex-start;
                    flex-wrap: wrap;
                }
            }
            """
        ),
    ),
    ui.output_ui("page"),
)


def server(input: Inputs, output: Outputs, session: Session):
    defaults = load_default_inputs()
    components_state: reactive.Value[pd.DataFrame] = reactive.Value(defaults["components"])
    assumptions_state: reactive.Value[pd.DataFrame] = reactive.Value(defaults["assumptions"])
    assessments_state: reactive.Value[pd.DataFrame] = reactive.Value(defaults["assessments"])
    result_state: reactive.Value[dict[str, object] | None] = reactive.Value(None)
    authenticated = reactive.Value(ALLOW_NO_PASSWORD or not APP_PASSWORD)

    def run_current_study() -> None:
        result_state.set(
            run_reserve_study(
                assumptions_state.get(),
                components_state.get(),
                assessments_state.get(),
                DEFAULT_PROJECTION_YEARS,
                DEFAULT_UNITS,
            )
        )

    @render.ui
    def page():
        if authenticated.get():
            return app_shell()
        return ui.div(
            ui.div(
                ui.h2("Ridge Park Reserve Study"),
                ui.input_password("password", "Password"),
                ui.input_action_button("login", "Enter", class_="rp-primary"),
                class_="rp-login",
            )
        )

    @reactive.effect
    @reactive.event(input.login)
    def _login():
        if input.password() == APP_PASSWORD:
            authenticated.set(True)
        else:
            ui.notification_show("Incorrect password.", type="error")

    @reactive.effect
    def _initial_run():
        if result_state.get() is None:
            run_current_study()

    @render.data_frame
    def components_grid():
        return render.DataGrid(
            display_components(components_state.get()),
            width="100%",
            height="640px",
            filters=True,
            editable=True,
        )

    @components_grid.set_patch_fn
    def _(*, patch: render.CellPatch):
        value = patch_frame(
            components_state,
            patch,
            display_labels=COMPONENT_REVERSE_LABELS,
            numeric_columns=NUMERIC_COMPONENT_COLUMNS,
        )
        components_state.set(raw_components(components_state.get()))
        return value

    @render.data_frame
    def assumptions_grid():
        return render.DataGrid(
            assumptions_state.get(),
            width="100%",
            height="360px",
            editable=True,
        )

    @assumptions_grid.set_patch_fn
    def _(*, patch: render.CellPatch):
        return patch_frame(assumptions_state, patch)

    @reactive.effect
    @reactive.event(input.reset_components)
    def _reset_components():
        components_state.set(defaults["components"])
        result_state.set(None)
        ui.notification_show("Components reset to source data.", type="message")

    @reactive.effect
    @reactive.event(input.run_study, input.run_study_assumptions)
    def _run_study():
        try:
            run_current_study()
            ui.notification_show("Reserve study updated.", type="message")
        except Exception as exc:
            ui.notification_show(str(exc), type="error", duration=10)

    @reactive.effect
    @reactive.event(input.open_upload)
    def _open_upload():
        ui.modal_show(
            ui.modal(
                ui.div(
                    ui.HTML(
                        'Upload a component schedule using <a href="#" onclick="document.getElementById(\'download_template\').click(); return false;">THIS TEMPLATE</a> in CSV, XLSX, XLSM, or XLSCSV format.'
                    ),
                    class_="rp-upload-copy",
                ),
                ui.input_file(
                    "component_upload",
                    None,
                    accept=[".csv", ".xlsx", ".xlsm", ".xls", ".xlscsv"],
                    multiple=False,
                    width="100%",
                    button_label="Upload",
                    placeholder="No file selected",
                ),
                title="Upload",
                easy_close=True,
                footer=ui.modal_button("Cancel"),
                size="m",
            )
        )

    @reactive.effect
    @reactive.event(input.component_upload)
    def _component_upload():
        uploaded = input.component_upload()
        if not uploaded:
            return
        try:
            file_info = uploaded[0]
            imported = read_component_upload(file_info["datapath"], file_info["name"])
            components_state.set(imported)
            result_state.set(None)
            ui.modal_remove()
            ui.notification_show(f"Imported {len(imported):,} components.", type="message")
        except Exception as exc:
            ui.notification_show(str(exc), type="error", duration=10)

    @render.download(filename="ridge_park_component_template.xlsx")
    def download_template():
        yield component_template_bytes()

    @render.download(filename="component_list_v2.csv")
    def download_components():
        yield components_state.get().to_csv(index=False)

    @render.table
    def statement_table():
        result = result_state.get()
        if result is None:
            return pd.DataFrame({"Metric": ["Run the study to see recommendations."], "Value": [""]})
        return result["statement_of_position_formatted"]

    @render.ui
    def results_summary():
        result = result_state.get()
        if result is None:
            return ui.p("Run the study to generate result tables and charts.")
        projection = result["reserve_projection"]
        final = projection.iloc[-1]
        return ui.div(
            ui.h4("Current Projection"),
            ui.p(
                f"Final year {int(final['year'])}: ending reserve balance {money(final['end_balance'])}."
            ),
        )

    def plot_at(index: int):
        result = result_state.get()
        if result is None:
            return None
        plots = build_all_plots(result)
        if index >= len(plots):
            return None
        return plots[index][1]

    @render.plot
    def plot_0():
        return plot_at(0)

    @render.plot
    def plot_1():
        return plot_at(1)

    @render.plot
    def plot_2():
        return plot_at(2)

    @render.plot
    def plot_3():
        return plot_at(3)

    @render.plot
    def plot_4():
        return plot_at(4)

    @render.plot
    def plot_5():
        return plot_at(5)


app = App(app_ui, server)
